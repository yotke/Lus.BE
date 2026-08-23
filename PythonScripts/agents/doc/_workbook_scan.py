# -*- coding: utf-8 -*-
"""Shared openpyxl helpers for document builder agents."""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any

from openpyxl.utils import get_column_letter

# Hebrew labels scanned in letterhead / header detection.
LETTERHEAD_LABELS = {
    "client": ("לקוח",),
    "planner": ("המתכנן", "המתכנן "),
    "account_number": ("מספר ח-ן", "מספר חשבון", "ח-ן"),
}

TABLE_HEADER_MARKERS = (
    "תאריך",
    "יום השבוע",
    "סהכ שעות",
    "מיקום",
    "נושא",
)

TOTALS_MARKERS = ('סה"כ', "סה''כ", "סהכ")


def sheet_names(wb) -> list[str]:
    return list(wb.sheetnames)


def pick_month_sheet(names: list[str]) -> str | None:
    """Prefer a March 2026 sheet; else the last non-rates/non-empty sheet."""
    for n in names:
        if "מרץ" in n and "2026" in n:
            return n
    skip = {"תעריפים", "גיליון2", "מקור"}
    candidates = [n for n in names if not any(s in n for s in skip) and n.strip()]
    return candidates[-1] if candidates else None


def column_widths(ws, max_col: int = 6) -> dict[str, float]:
    out: dict[str, float] = {}
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        w = ws.column_dimensions[letter].width
        if w is not None:
            out[letter] = float(w)
    return out


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def find_row_with_label(ws, labels: tuple[str, ...], max_row: int = 20) -> int | None:
    for r in range(1, max_row + 1):
        for c in range(1, 8):
            t = cell_text(ws.cell(r, c).value)
            if not t:
                continue
            for label in labels:
                if label in t:
                    return r
    return None


def find_table_header_row(ws, max_row: int = 30) -> int | None:
    for r in range(1, max_row + 1):
        row_text = " ".join(cell_text(ws.cell(r, c).value) for c in range(1, 6))
        hits = sum(1 for m in TABLE_HEADER_MARKERS if m in row_text)
        if hits >= 3:
            return r
    return None


def find_totals_row(ws, start_row: int, max_row: int = 80) -> int | None:
    for r in range(start_row, max_row + 1):
        a = cell_text(ws.cell(r, 1).value)
        if any(m in a for m in TOTALS_MARKERS) and "שעות" not in a:
            return r
    return None


def find_billing_start(ws, totals_row: int, max_row: int = 90) -> int | None:
    for r in range(totals_row + 1, max_row + 1):
        a = cell_text(ws.cell(r, 1).value)
        if "שעות עבודה" in a or "מחיר" in a:
            return r
    return None


def find_declaration_start(ws, billing_start: int, max_row: int = 120) -> int | None:
    for r in range(billing_start, max_row + 1):
        a = cell_text(ws.cell(r, 1).value)
        if "הצהר" in a or "חתימ" in a:
            return r
    return None


def scan_letterhead(ws, max_row: int = 12) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, labels in LETTERHEAD_LABELS.items():
        row = find_row_with_label(ws, labels, max_row=max_row)
        if row is None:
            continue
        for c in range(1, 8):
            t = cell_text(ws.cell(row, c).value)
            if any(l in t for l in labels):
                # value is typically the next populated cell on the row
                for c2 in range(c + 1, 8):
                    v = ws.cell(row, c2).value
                    if v is not None and cell_text(v):
                        out[key] = cell_text(v)
                        break
                break
    # Account number often lives in column E with the label embedded
    for r in range(1, max_row + 1):
        for c in range(1, 8):
            t = cell_text(ws.cell(r, c).value)
            if "ח-ן" in t or "חשבון" in t:
                m = re.search(r"(\d{6,})", t)
                if m:
                    out.setdefault("account_number", m.group(1))
    return out


def merged_range_count(ws) -> int:
    return len(ws.merged_cells.ranges)


def title_row(ws) -> int | None:
    for r in range(1, 8):
        t = cell_text(ws.cell(r, 1).value)
        if "דוח ביצוע שעות" in t:
            return r
    return None


# ── Learning from the exemplar's CONTENT, not just its geometry ────────────────────
# A workbook the user hands over is a year of their own decisions: which sites they
# work at, how they word a task, what they charge, how many hours a visit runs. Reading
# that turns the interview from generic prompts into their own vocabulary, and gives the
# billing questions answers grounded in what they actually billed last month.


def _numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def scan_month_content(ws, header_row: int, totals_row: int) -> dict[str, Any]:
    """Locations, subjects and hour values from one month's data band."""
    locations: list[str] = []
    subjects: list[str] = []
    hours: list[float] = []

    for r in range(header_row + 1, totals_row):
        hours_value = _numeric(ws.cell(r, 3).value)
        if hours_value is not None:
            hours.append(hours_value)

        location = cell_text(ws.cell(r, 4).value)
        if location and location not in locations:
            locations.append(location)

        subject = cell_text(ws.cell(r, 5).value)
        if subject:
            subjects.append(subject)

    return {"locations": locations, "subjects": subjects, "hours": hours}


def scan_billing(ws, billing_start: int, declaration_start: int) -> dict[str, Any]:
    """
    The money block: the rate the user charges and the VAT they apply.

    Label and value are read as a pair, scanning the row for the first number, because the
    block migrates between columns across the exemplar's own sheets (column C in the oldest
    sheet, column D in every later one) — hard-coding a column is how the formulas got lost
    in the first place.
    """
    out: dict[str, Any] = {}

    for r in range(billing_start, declaration_start):
        label = cell_text(ws.cell(r, 1).value)
        if not label:
            continue

        value = None
        for c in range(2, 9):
            value = _numeric(ws.cell(r, c).value)
            if value is not None:
                break

        if "מחיר לשעת" in label and value:
            out["rate"] = value
        elif 'מע"מ' in label or "מעמ" in label:
            # The percentage usually lives in the LABEL ("מע\"מ 18%"), not the cell.
            match = re.search(r"(\d+(?:\.\d+)?)\s*%", label)
            if match:
                out["vat_percent"] = float(match.group(1))

    return out


def scan_totals_row(ws, totals_row: int) -> dict[str, Any]:
    """Hours / carried-in / remaining from a month's totals band."""
    out: dict[str, Any] = {}
    hours = _numeric(ws.cell(totals_row, 3).value)
    carry_in = _numeric(ws.cell(totals_row, 4).value)
    remaining = _numeric(ws.cell(totals_row, 5).value)
    if hours is not None:
        out["hours"] = hours
    if carry_in is not None:
        out["carry_in"] = carry_in
    if remaining is not None:
        out["remaining"] = remaining
    return out
