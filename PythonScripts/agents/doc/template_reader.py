# -*- coding: utf-8 -*-
"""Importer: exemplar .xlsx → DocumentTemplate skeleton (no LLM)."""

from __future__ import annotations

import os

from openpyxl import load_workbook

from agents.doc._workbook_scan import (
    cell_text,
    column_widths,
    scan_billing,
    scan_month_content,
    scan_totals_row,
    find_billing_start,
    find_declaration_start,
    find_table_header_row,
    find_totals_row,
    merged_range_count,
    pick_month_sheet,
    scan_letterhead,
    sheet_names,
    title_row,
)


def _resolve_path(agent_input: dict) -> str | None:
    if not agent_input:
        return None
    for key in ("FilePath", "filePath", "Path", "path"):
        if agent_input.get(key):
            return str(agent_input[key])
    return None


def _learn_from_history(wb, names, current_sheet):
    """
    Read the user's OWN past months, not just the sheet we are templating.

    A workbook handed over is a year of decisions: the sites they visit, how they word a
    task, what they charge, how long a visit runs. Learning those turns the interview from
    generic prompts into the user's own vocabulary, and gives the billing questions answers
    grounded in what they actually billed — instead of the invented defaults that were
    hard-coded into the planner.
    """
    locations: list[str] = []
    subjects: list[str] = []
    hours: list[float] = []
    rates: list[float] = []
    vat_percent = None
    last_remaining = None
    day_totals: list[float] = []

    for name in names:
        try:
            sheet = wb[name]
            header = find_table_header_row(sheet)
            if header is None:
                continue
            totals = find_totals_row(sheet, header + 1)
            if totals is None:
                continue

            content = scan_month_content(sheet, header, totals)
            for location in content["locations"]:
                if location not in locations:
                    locations.append(location)
            subjects.extend(content["subjects"])
            hours.extend(content["hours"])

            # Hours grouped by the date they hang under: the exemplar writes the date once
            # and leaves it blank for the following segments of the same day.
            current_day = 0.0
            for r in range(header + 1, totals):
                if sheet.cell(r, 1).value is not None:
                    if current_day:
                        day_totals.append(current_day)
                    current_day = 0.0
                value = sheet.cell(r, 3).value
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    current_day += float(value)
            if current_day:
                day_totals.append(current_day)

            billing_start = find_billing_start(sheet, totals)
            declaration_start = find_declaration_start(sheet, billing_start or totals + 3)
            if billing_start:
                billing = scan_billing(sheet, billing_start, declaration_start or billing_start + 6)
                if billing.get("rate"):
                    rates.append(billing["rate"])
                if billing.get("vat_percent") and vat_percent is None:
                    vat_percent = billing["vat_percent"]

            # The most recent month that is not the one being templated carries the balance
            # this document should start from (the exemplar's 810 -> 760 -> 728 chain).
            if name != current_sheet:
                band = scan_totals_row(sheet, totals)
                if band.get("remaining") is not None:
                    last_remaining = band["remaining"]
        except Exception:
            # One unreadable sheet must never cost the whole import.
            continue

    # Subjects repeat across months; the ones that repeat are the ones worth suggesting.
    counts: dict[str, int] = {}
    for subject in subjects:
        key = subject.strip()
        if key:
            counts[key] = counts.get(key, 0) + 1
    common_subjects = [s for s, _ in sorted(counts.items(), key=lambda kv: -kv[1])[:8]]

    positive_hours = [h for h in hours if h > 0]

    return {
        "Locations": locations[:8],
        "Subjects": common_subjects,
        "Rates": sorted({round(r, 2) for r in rates}, reverse=True)[:4],
        "VatPercent": vat_percent,
        "LastRemaining": last_remaining,
        # Plausibility bounds, learned rather than invented: a value outside the range this
        # user has ever billed is worth questioning (a 115-hour row is a typo, not a day).
        "MinHours": min(positive_hours) if positive_hours else None,
        "MaxHours": max(positive_hours) if positive_hours else None,
        "MaxHoursPerDay": max(day_totals) if day_totals else None,
        "MonthsSeen": len(names),
    }


def run(*, draft, agent_input, lang="he"):
    path = _resolve_path(agent_input)
    if not path or not os.path.isfile(path):
        raise FileNotFoundError(f"Exemplar file not found: {path!r}")

    wb = load_workbook(path, data_only=False, read_only=False)
    try:
        names = sheet_names(wb)
        sheet_name = pick_month_sheet(names)
        if not sheet_name:
            raise ValueError("No month sheet found in workbook")

        ws = wb[sheet_name]
        header_row = find_table_header_row(ws)
        if header_row is None:
            raise ValueError("Table header row not found")

        data_start = header_row + 1
        totals_row = find_totals_row(ws, data_start) or (data_start + 14)
        billing_start = find_billing_start(ws, totals_row) or (totals_row + 3)
        declaration_start = find_declaration_start(ws, billing_start) or (billing_start + 6)
        t_row = title_row(ws) or 4
        letterhead = scan_letterhead(ws)
        merge_count = merged_range_count(ws)
        rtl = bool(getattr(ws.sheet_view, "rightToLeft", False))

        # Column headings exactly as the workbook writes them — the canvas shows the
        # user's own labels rather than ours.
        headers = []
        for c in range(1, ws.max_column + 1):
            label = cell_text(ws.cell(header_row, c).value)
            if label:
                headers.append(label)

        title_text = cell_text(ws.cell(t_row, 1).value)
        org_name = ""
        for r in range(1, t_row):
            candidate = cell_text(ws.cell(r, 1).value)
            if candidate:
                org_name = candidate
                break

        # The billing block's own wording (rate / subtotal / VAT / total), so the preview
        # reproduces the document's vocabulary instead of inventing labels.
        billing_labels = []
        for r in range(billing_start, declaration_start):
            label = cell_text(ws.cell(r, 1).value)
            if label:
                billing_labels.append(label)

        declaration_text = cell_text(ws.cell(declaration_start, 1).value)

        # What the workbook TEACHES, beyond where its cells are.
        #
        # Read from a data_only handle: the totals band is formulas (=SUM(C11:C24),
        # =+'previous month'!E31), and a formula-valued workbook hands back the formula
        # TEXT, so the carried-in balance would come back empty. data_only=True returns the
        # values Excel last calculated.
        values_wb = load_workbook(path, data_only=True, read_only=False)
        try:
            learned = _learn_from_history(values_wb, names, sheet_name)
        finally:
            values_wb.close()

        five_blocks = {
            "title": t_row,
            "letterhead": max(5, t_row + 1),
            "tableHeader": header_row,
            "dataBandStart": data_start,
            "totals": totals_row,
            "billing": billing_start,
            "declaration": declaration_start,
        }

        return {
            "SheetName": sheet_name,
            "Rtl": rtl,
            "ColumnWidths": column_widths(ws),
            "MergeCount": merge_count,
            "MergePolicy": "group-date-columns-AB",
            "DataBandStartRow": data_start,
            "TableHeaderRow": header_row,
            "TitleRow": t_row,
            "TotalsRow": totals_row,
            "BillingStartRow": billing_start,
            "DeclarationStartRow": declaration_start,
            "Letterhead": letterhead,
            "FiveBlocks": five_blocks,
            # Everything the canvas needs to PREVIEW the finished document, not just to
            # know where the data band starts: the title band, the letterhead lines, the
            # real column headings and the billing block's own labels. Geometry alone
            # renders a bare grid; these are what make it look like the workbook.
            "Patches": [
                {"Op": "SetField", "Path": "template.rtl", "Value": rtl},
                {"Op": "SetField", "Path": "template.dataBandStartRow", "Value": data_start},
                {"Op": "SetField", "Path": "template.mergePolicy", "Value": "group-date-columns-AB"},
                {"Op": "SetField", "Path": "template.sheetName", "Value": sheet_name},
                {"Op": "SetField", "Path": "template.tableHeaderRow", "Value": header_row},
                {"Op": "SetField", "Path": "template.titleRow", "Value": t_row},
                {"Op": "SetField", "Path": "template.totalsRow", "Value": totals_row},
                {"Op": "SetField", "Path": "template.billingStartRow", "Value": billing_start},
                {"Op": "SetField", "Path": "template.declarationStartRow", "Value": declaration_start},
                {"Op": "SetField", "Path": "template.mergeCount", "Value": merge_count},
                {"Op": "SetField", "Path": "template.columnWidths", "Value": column_widths(ws)},
                {"Op": "SetField", "Path": "template.headers", "Value": headers},
                {"Op": "SetField", "Path": "template.title", "Value": title_text},
                {"Op": "SetField", "Path": "template.orgName", "Value": org_name},
                {"Op": "SetField", "Path": "template.plannerName", "Value": letterhead.get("planner")},
                {"Op": "SetField", "Path": "template.clientName", "Value": letterhead.get("client")},
                {"Op": "SetField", "Path": "template.billingLabels", "Value": billing_labels},
                {"Op": "SetField", "Path": "template.declarationText", "Value": declaration_text},
                {"Op": "SetField", "Path": "template.learned", "Value": learned},
            ]
            + (
                [{"Op": "SetField", "Path": "accountNumber", "Value": letterhead["account_number"]}]
                if letterhead.get("account_number")
                else []
            ),
        }
    finally:
        wb.close()
